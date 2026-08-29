from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import send_file
from datetime import datetime
from io import BytesIO
import sqlite3


def export_excel(db_path):

    # ==============================
    # CONNECT TO DATABASE
    # ==============================
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM predictions ORDER BY id DESC")
    rows = cursor.fetchall()

    columns = [desc[0] for desc in cursor.description]

    # ==============================
    # HUMAN READABLE MAPS
    # ==============================
    region_map = {
        0: "Asia",
        1: "Europe",
        2: "North America",
        3: "South America",
        4: "Africa"
    }

    transport_map = {
        0: "Truck",
        1: "Rail",
        2: "Ship",
        3: "Air"
    }

    weather_map = {
        0: "Clear",
        1: "Rain",
        2: "Storm",
        3: "Snow",
        4: "Fog"
    }

    demand_map = {
        0: "Low",
        1: "Medium",
        2: "High"
    }

    traffic_map = {
        0: "Low",
        1: "Medium",
        2: "High"
    }

    port_map = {
        0: "No",
        1: "Yes"
    }

    # ==============================
    # CREATE WORKBOOK
    # ==============================
    wb = Workbook()
    ws = wb.active
    ws.title = "Supply Chain Report"

    # ==============================
    # HEADER STYLE
    # ==============================
    header_fill = PatternFill(
        start_color="1E3A8A",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ==============================
    # CLEAN COLUMN NAMES
    # ==============================
    clean_columns = [
        "ID",
        "Date",
        "Supplier",
        "Region",
        "Transport",
        "Delay (Days)",
        "Weather",
        "Demand",
        "Inventory",
        "Traffic",
        "Port Delay",
        "Order Value ($)",
        "Fuel Cost ($/unit)",
        "Risk"
    ]

    # ==============================
    # HEADER
    # ==============================
    for col_num, col_name in enumerate(clean_columns, 1):

        cell = ws.cell(
            row=1,
            column=col_num,
            value=col_name
        )

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # ==============================
    # DATA
    # ==============================
    for row_num, db_row in enumerate(rows, 2):

        row = list(db_row)

        # Convert values to human-readable format
        row[3] = region_map.get(row[3], "Unknown")
        row[4] = transport_map.get(row[4], "Unknown")
        row[6] = weather_map.get(row[6], "Unknown")
        row[7] = demand_map.get(row[7], "Unknown")
        row[9] = traffic_map.get(row[9], "Unknown")
        row[10] = port_map.get(row[10], "No")

        for col_num, value in enumerate(row, 1):

            cell = ws.cell(
                row=row_num,
                column=col_num,
                value=value
            )

            cell.alignment = Alignment(horizontal="center")
            cell.border = border

            # ==============================
            # RISK COLOR
            # ==============================
            if columns[col_num - 1] == "risk":

                if value == "High":

                    cell.fill = PatternFill(
                        start_color="FECACA",
                        fill_type="solid"
                    )

                elif value == "Medium":

                    cell.fill = PatternFill(
                        start_color="FED7AA",
                        fill_type="solid"
                    )

                elif value == "Low":

                    cell.fill = PatternFill(
                        start_color="BBF7D0",
                        fill_type="solid"
                    )

    # ==============================
    # AUTO COLUMN WIDTH
    # ==============================
    for column in ws.columns:

        max_length = 0

        for cell in column:

            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[
            column[0].column_letter
        ].width = max_length + 4

    # ==============================
    # SAVE EXCEL IN MEMORY
    # ==============================
    output = BytesIO()

    wb.save(output)

    output.seek(0)

    # ==============================
    # CLOSE DATABASE
    # ==============================
    conn.close()

    # ==============================
    # GENERATE FILENAME
    # ==============================
    filename = (
        "supply_chain_report_"
        + datetime.now().strftime("%Y%m%d_%H%M%S")
        + ".xlsx"
    )

    # ==============================
    # DOWNLOAD DIRECTLY
    # ==============================
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )